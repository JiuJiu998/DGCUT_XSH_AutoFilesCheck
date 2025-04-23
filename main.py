import os
import re
import shutil
import pandas as pd
import zipfile
import rarfile
import py7zr
import logging
from pyunpack import Archive
from pathlib import Path
from openpyxl import load_workbook
from datetime import datetime

# 配置路径
INPUT_DIR = Path('./input')
OUTPUT_DOC = Path('./output/doc')
OUTPUT_IMG = Path('./output/image')
ERROR_DIR = Path('./error')
TEMP_DIR = Path('./temp')
LOG_FILE = Path('./run_log.txt')
SUMMARY_FILE = Path('./output/汇总表.xlsx')

# 日志配置：同时打印到控制台和文件
logger = logging.getLogger()
logger.setLevel(logging.INFO)

formatter = logging.Formatter('%(asctime)s - %(levelname)s - %(message)s')

# 文件输出
file_handler = logging.FileHandler(LOG_FILE, mode='a', encoding='utf-8')

file_handler.setFormatter(formatter)
logger.addHandler(file_handler)

# 控制台输出
console_handler = logging.StreamHandler()
console_handler.setFormatter(formatter)
logger.addHandler(console_handler)


# 自动创建目录
for path in [INPUT_DIR, OUTPUT_DOC, OUTPUT_IMG, ERROR_DIR, TEMP_DIR, SUMMARY_FILE.parent]:
    path.mkdir(parents=True, exist_ok=True)

# 清理文件名非法字符
def clean_filename(name):
    return re.sub(r'[\\/:*?"<>|\n\r\t]', '_', name)

# 自动重命名函数
def get_unique_path(base_path):
    if not base_path.exists():
        return base_path
    i = 1
    while True:
        new_path = base_path.with_name(f"{base_path.stem}_{i}{base_path.suffix}")
        if not new_path.exists():
            return new_path
        i += 1

# 解压函数
def extract_archive(file_path, extract_to):
    try:
        Archive(str(file_path)).extractall(str(extract_to))
        return True
    except Exception as e:
        logging.error(f"解压失败：{file_path} - {e}")
        return False

# 读取Excel数据
def read_excel_info(file_path):
    try:
        wb = load_workbook(file_path)
        ws = wb.active
        values = [ws[f"{col}8"].value for col in "ABCDEFG"]
        if None in values or len(values) != 7:
            return None
        return values
    except Exception as e:
        logging.warning(f"读取Excel失败：{file_path} - {e}")
        return None

# 搜索文件
def find_file_by_ext(root, exts):
    matches = []
    for ext in exts:
        matches.extend(root.rglob(f'*{ext}'))
    return matches

# 主处理函数
def process_archives():
    summary_data = []

    for archive in INPUT_DIR.iterdir():
        if archive.suffix.lower() not in ['.zip', '.rar', '.7z']:
            continue

        logging.info(f"处理压缩包：{archive.name}")
        shutil.rmtree(TEMP_DIR, ignore_errors=True)
        TEMP_DIR.mkdir(exist_ok=True)

        success = extract_archive(archive, TEMP_DIR)
        if not success:
            shutil.move(str(archive), ERROR_DIR / archive.name)
            continue

        excel_files = find_file_by_ext(TEMP_DIR, ['.xlsx', '.xls'])
        if not excel_files:
            logging.warning(f"未找到Excel文件：{archive.name}")
            shutil.move(str(archive), ERROR_DIR / archive.name)
            continue

        info = read_excel_info(excel_files[0])
        if not info:
            logging.warning(f"Excel数据格式错误：{archive.name}")
            shutil.move(str(archive), ERROR_DIR / archive.name)
            continue

        category, name, gender, college, clazz, student_id, phone = info
        remark = []

        base_name = clean_filename(f"{name}-{student_id}-{category}")

        # 保存word文档
        word_files = find_file_by_ext(TEMP_DIR, ['.docx', '.doc'])
        if word_files:
            new_name = get_unique_path(OUTPUT_DOC / f"{base_name}.docx")
            shutil.copy(word_files[0], new_name)
        else:
            remark.append("缺少doc文件")

        # 保存图片
        image_files = find_file_by_ext(TEMP_DIR, ['.jpg', '.jpeg', '.png'])
        if image_files:
            img_ext = image_files[0].suffix.lower()
            new_name = get_unique_path(OUTPUT_IMG / f"{base_name}{img_ext}")
            shutil.copy(image_files[0], new_name)
        else:
            remark.append("缺少图片")

        summary_data.append({
            "类别": category,
            "姓名": name,
            "性别": gender,
            "学院": college,
            "班级": clazz,
            "学号": student_id,
            "联系电话": phone,
            "备注": "正常" if not remark else "、".join(remark)
        })

        logging.info(f"成功处理：{archive.name} - {base_name}")

    # 保存汇总表
    df = pd.DataFrame(summary_data)
    df.to_excel(SUMMARY_FILE, index=False)
    logging.info(f"汇总完成，共计：{len(summary_data)} 人，已保存至 {SUMMARY_FILE}")

if __name__ == '__main__':
    logging.info("===== 开始处理 =====")
    process_archives()
    logging.info("===== 处理结束 =====")
